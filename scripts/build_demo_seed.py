import openpyxl
import json
import hashlib

wb = openpyxl.load_workbook('d:/event-tracker/EventTrack_Demo_Data.xlsx')

data = {}
for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    headers = [str(h).strip() if h is not None else f'col_{i}' for i, h in enumerate(rows[0])]
    sheet_data = []
    for r in rows[1:]:
        if any(cell is not None for cell in r):
            row_dict = {}
            for h, cell in zip(headers, r):
                row_dict[h] = str(cell).strip() if cell is not None else None
            sheet_data.append(row_dict)
    data[sheet] = sheet_data

participants = data.get('Participants', [])
organizers = data.get('Organizers', [])
volunteers = data.get('Volunteers', [])
events = data.get('Events', [])
registrations = data.get('Registrations', [])
attendance = data.get('Attendance', [])

# Map Attendance lookup: registration_id -> status ('Present' / 'Absent')
attendance_map = {a['registration_id']: a['attendance'] for a in attendance}

# Create unified users dataset
users_list = []

def make_uuid(prefix_digit: str, num_raw: str) -> str:
    clean_num = ''.join(c for c in str(num_raw) if c.isdigit()) or '1'
    suffix = f"{prefix_digit}{clean_num.zfill(11)}"
    return f"00000000-0000-4000-8000-{suffix}"

# Organizers (O001..O005) -> prefix 1
for o in organizers:
    users_list.append({
        'id': make_uuid('1', o['organizer_id']),
        'legacy_id': o['organizer_id'],
        'email': o['email'],
        'full_name': o['name'],
        'role': 'organizer',
        'usn': None,
        'branch': None,
        'phone': None
    })

# Volunteers (V001..V005) -> prefix 2
for v in volunteers:
    users_list.append({
        'id': make_uuid('2', v['volunteer_id']),
        'legacy_id': v['volunteer_id'],
        'email': v['email'],
        'full_name': v['name'],
        'role': 'volunteer',
        'usn': None,
        'branch': None,
        'phone': None
    })

# Participants (P001..P100) -> prefix 3
for p in participants:
    users_list.append({
        'id': make_uuid('3', p['participant_id']),
        'legacy_id': p['participant_id'],
        'email': p['email'],
        'full_name': p['name'],
        'role': 'participant',
        'usn': p.get('usn'),
        'branch': p.get('branch'),
        'phone': p.get('phone')
    })

# User lookup by legacy_id
user_by_legacy = {u['legacy_id']: u for u in users_list}
user_by_email = {u['email'].lower(): u for u in users_list}

# Process Events -> prefix 4
processed_events = []
event_by_legacy = {}
for e in events:
    org_user = user_by_legacy.get(e['organizer_id'])
    org_id = org_user['id'] if org_user else users_list[0]['id']
    
    # Standardize start_time/end_time from date
    event_date = e['date'] # e.g. 2026-08-06
    start_time = f"{event_date}T09:00:00.000Z"
    end_time = f"{event_date}T17:00:00.000Z"
    
    ev_obj = {
        'id': make_uuid('4', e['event_id']),
        'legacy_id': e['event_id'],
        'title': e['event_name'],
        'description': f"Official campus event: {e['event_name']}. Join talks, interactive sessions, and practical track activities.",
        'location': 'Main Campus',
        'venue': e['venue'],
        'start_time': start_time,
        'end_time': end_time,
        'capacity': int(e['capacity']) if e.get('capacity') else 200,
        'organizer_id': org_id,
        'organizer_name': org_user['full_name'] if org_user else 'Organizer',
        'status': 'published',
        'is_registration_open': True
    }
    processed_events.append(ev_obj)
    event_by_legacy[e['event_id']] = ev_obj

# Process Registrations & Attendance -> prefix 5 (reg) & 6 (att)
processed_registrations = []
processed_attendance = []

for reg in registrations:
    r_id = reg['registration_id']
    p_user = user_by_legacy.get(reg['participant_id'])
    ev_item = event_by_legacy.get(reg['event_id'])
    
    if not p_user or not ev_item:
        continue
        
    att_status = attendance_map.get(r_id, 'Absent')
    is_checked_in = att_status.strip().lower() == 'present'
    
    # Generate deterministic QR Token
    token_seed = f"eventtrack_qr_{r_id}_{p_user['id']}_{ev_item['id']}"
    qr_token = hashlib.sha256(token_seed.encode('utf-8')).hexdigest()[:32]
    
    reg_uuid = make_uuid('5', r_id)
    
    reg_obj = {
        'id': reg_uuid,
        'legacy_id': r_id,
        'event_id': ev_item['id'],
        'participant_id': p_user['id'],
        'qr_token': qr_token,
        'qr_version': 1,
        'status': 'checked_in' if is_checked_in else 'registered',
        'registered_at': ev_item['start_time'], # or date
        'cancelled_at': None,
        'events': ev_item,
        'users': p_user
    }
    processed_registrations.append(reg_obj)
    
    if is_checked_in:
        vol_user = user_by_legacy.get('V001') # Default Volunteer V001
        att_uuid = make_uuid('6', r_id)
        processed_attendance.append({
            'id': att_uuid,
            'registration_id': reg_uuid,
            'event_id': ev_item['id'],
            'participant_id': p_user['id'],
            'checked_in_by': vol_user['id'] if vol_user else p_user['id'],
            'checked_in_at': ev_item['start_time'],
            'method': 'qr_scan'
        })

demo_dataset = {
    'users': users_list,
    'events': processed_events,
    'registrations': processed_registrations,
    'attendance': processed_attendance
}

import os

os.makedirs('d:/event-tracker/src/data', exist_ok=True)
os.makedirs('d:/event-tracker/supabase', exist_ok=True)

# Save src/data/demo-dataset.json
with open('d:/event-tracker/src/data/demo-dataset.json', 'w') as f:
    json.dump(demo_dataset, f, indent=2)

# Generate SQL Seed File
all_emails_sql = "', '".join([u['email'] for u in users_list])

sql_lines = [
    "-- =============================================================================",
    "-- EventTrack - Seed Data imported from EventTrack_Demo_Data.xlsx",
    "-- All users default password: 12345678",
    "-- =============================================================================",
    "CREATE EXTENSION IF NOT EXISTS pgcrypto;",
    "",
    "-- Delete existing demo records to prevent duplicate key conflicts",
    f"DELETE FROM auth.users WHERE email IN ('{all_emails_sql}');",
    f"DELETE FROM public.users WHERE email IN ('{all_emails_sql}');",
    ""
]

# Auth Users & Public Users
for u in users_list:
    meta = json.dumps({
        'full_name': u['full_name'],
        'role': u['role'],
        'usn': u['usn'],
        'branch': u['branch'],
        'phone': u['phone']
    })
    
    sql_lines.append(f"""
INSERT INTO auth.users (
  id, instance_id, email, encrypted_password, email_confirmed_at, raw_app_meta_data, raw_user_meta_data, created_at, updated_at, role, aud
) VALUES (
  '{u['id']}',
  '00000000-0000-0000-0000-000000000000',
  '{u['email']}',
  crypt('12345678', gen_salt('bf')),
  now(),
  '{{"provider":"email","providers":["email"]}}',
  '{meta}',
  now(),
  now(),
  'authenticated',
  'authenticated'
) ON CONFLICT (id) DO UPDATE SET encrypted_password = crypt('12345678', gen_salt('bf'));

INSERT INTO public.users (
  id, email, full_name, usn, branch, phone, role
) VALUES (
  '{u['id']}',
  '{u['email']}',
  '{u['full_name'].replace("'", "''")}',
  {f"'{u['usn']}'" if u['usn'] else 'NULL'},
  {f"'{u['branch']}'" if u['branch'] else 'NULL'},
  {f"'{u['phone']}'" if u['phone'] else 'NULL'},
  '{u['role']}'
) ON CONFLICT (id) DO UPDATE SET
  full_name = EXCLUDED.full_name,
  usn = EXCLUDED.usn,
  branch = EXCLUDED.branch,
  role = EXCLUDED.role;
""")

# Events
for e in processed_events:
    sql_lines.append(f"""
INSERT INTO public.events (
  id, title, description, location, venue, start_time, end_time, capacity, organizer_id, status, is_registration_open
) VALUES (
  '{e['id']}',
  '{e['title'].replace("'", "''")}',
  '{e['description'].replace("'", "''")}',
  '{e['location']}',
  '{e['venue']}',
  '{e['start_time']}',
  '{e['end_time']}',
  {e['capacity']},
  '{e['organizer_id']}',
  '{e['status']}',
  {str(e['is_registration_open']).upper()}
) ON CONFLICT (id) DO NOTHING;
""")

# Registrations
for r in processed_registrations:
    sql_lines.append(f"""
INSERT INTO public.registrations (
  id, event_id, participant_id, qr_token, qr_version, status, registered_at
) VALUES (
  '{r['id']}',
  '{r['event_id']}',
  '{r['participant_id']}',
  '{r['qr_token']}',
  1,
  '{r['status']}',
  '{r['registered_at']}'
) ON CONFLICT (id) DO UPDATE SET status = EXCLUDED.status;
""")

# Attendance
for a in processed_attendance:
    sql_lines.append(f"""
INSERT INTO public.attendance (
  id, registration_id, event_id, participant_id, checked_in_by, checked_in_at, method
) VALUES (
  '{a['id']}',
  '{a['registration_id']}',
  '{a['event_id']}',
  '{a['participant_id']}',
  '{a['checked_in_by']}',
  '{a['checked_in_at']}',
  '{a['method']}'
) ON CONFLICT (registration_id) DO NOTHING;
""")

with open('d:/event-tracker/supabase/seed_excel_data.sql', 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_lines))

print("Generated supabase/seed_excel_data.sql successfully!")

print(f"Generated demo-dataset.json successfully!")
print(f"Total Users: {len(users_list)} (100 Participants, 5 Organizers, 5 Volunteers)")
print(f"Total Events: {len(processed_events)}")
print(f"Total Registrations: {len(processed_registrations)}")
print(f"Total Attendance (Checked-in): {len(processed_attendance)}")
